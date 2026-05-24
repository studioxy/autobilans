from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from autobilans.models import ZoisRow


def _to_float(value: object) -> float:
    if value in (None, ""):
        return 0.0
    return float(value)


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def _header_index_map(header_row: tuple[object, ...]) -> dict[str, int]:
    return {
        _normalize_header(value): index
        for index, value in enumerate(header_row)
        if _normalize_header(value)
    }


def _row_value(
    row: tuple[object, ...],
    header_map: dict[str, int],
    aliases: tuple[str, ...],
    fallback_index: int | None = None,
) -> object:
    for alias in aliases:
        index = header_map.get(alias)
        if index is not None and index < len(row):
            return row[index]
    if fallback_index is not None and fallback_index < len(row):
        return row[fallback_index]
    return None


def parse_zois(path: str | Path, *, company: str | None = None, year: int | None = None) -> list[ZoisRow]:
    workbook_path = Path(path)
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = _header_index_map(header_row)

    rows: list[ZoisRow] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        account_no = _row_value(row, header_map, ("numer",), 0)
        if account_no in (None, ""):
            continue

        balance_code = _row_value(row, header_map, ("s_12_1",), 12)
        additional_balance_codes = tuple(
            str(value).strip()
            for value in (
                _row_value(row, header_map, ("s_12_2",), 13),
                _row_value(row, header_map, ("s_12_3",), 14),
            )
            if value not in (None, "")
        )
        rows.append(
            ZoisRow(
                account_no=str(account_no).strip(),
                name=str(_row_value(row, header_map, ("nazwa",), 1) or "").strip(),
                name_2=str(_row_value(row, header_map, ("nazwa 2",), None) or "").strip(),
                closing_debit=_to_float(_row_value(row, header_map, ("saldo wn",), 9)),
                closing_credit=_to_float(_row_value(row, header_map, ("saldo ma",), 10)),
                persaldo=_to_float(_row_value(row, header_map, ("persaldo",), 11)),
                balance_code=str(balance_code).strip() if balance_code not in (None, "") else None,
                additional_balance_codes=additional_balance_codes,
                company=company,
                year=year,
                source_path=str(workbook_path),
            )
        )

    return rows
