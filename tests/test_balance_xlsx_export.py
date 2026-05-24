from pathlib import Path

from openpyxl import load_workbook

from autobilans.exporters import export_balance_xlsx
from autobilans.validation.compare import ValidationItem


def test_export_balance_xlsx_writes_file_with_summary(tmp_path: Path) -> None:
    target = export_balance_xlsx(
        calculated_balance={
            "Aktywa_A_I": 100.0,
            "Aktywa_A_I__POS": 100.0,
            "Pasywa_A_I": 100.0,
        },
        validation_results=[
            ValidationItem(code="Aktywa_A_I", expected=100.0, actual=100.0, delta=0.0, matched=True),
            ValidationItem(code="Pasywa_A_I", expected=90.0, actual=100.0, delta=10.0, matched=False),
        ],
        output_dir=tmp_path,
    )

    assert target.exists()

    wb = load_workbook(target)
    assert "Bilans" in wb.sheetnames
    assert "Podsumowanie" in wb.sheetnames
    ws = wb["Bilans"]
    assert ws["A1"].value == "Kod"
    assert ws["A2"].value == "Aktywa_A_I"
    assert ws["A3"].value == "Pasywa_A_I"
    # Technical __POS bucket should be excluded from sheet rows.
    assert ws["A4"].value is None
